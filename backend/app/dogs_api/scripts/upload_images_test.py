import os
from pathlib import Path

import openpyxl
from openpyxl_image_loader import SheetImageLoader


def extraer_imagenes_excel(archivo_excel, carpeta_salida='imagenes_perros', fila_inicio=4, columna_imagen=1):
    Path(carpeta_salida).mkdir(parents=True, exist_ok=True)
    print(f"📂 Cargando archivo: {archivo_excel}")
    wb = openpyxl.load_workbook(archivo_excel)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet)

    imagenes_guardadas = []
    imagenes_faltantes = []
    perro_id = 1
    errores = []

    print(f"🔍 Extrayendo imagenes desde fila {fila_inicio}...\n")

    for row in range(fila_inicio, sheet.max_row + 1):
        # Buscar imagen en cualquier columna de esta fila
        imagen_encontrada = False
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            if image_loader.image_in(cell.coordinate):
                try:
                    # Obtener la imagen
                    img = image_loader.get(cell.coordinate)
                    # Crear nombre con ID secuencial
                    nombre_archivo = f"perro_id_{perro_id}.png"
                    ruta_completa = os.path.join(carpeta_salida, nombre_archivo)
                    # Guardar imagen
                    img.save(ruta_completa)
                    # Guardar información
                    info = {
                        'id': perro_id,
                        'fila': row,
                        'columna': col,
                        'celda': cell.coordinate,
                        'archivo': nombre_archivo,
                        'ruta': ruta_completa
                    }
                    imagenes_guardadas.append(info)
                    print(f"✅ ID {perro_id:3d} | Fila {row:3d} | Celda {cell.coordinate:4s} | {nombre_archivo}")

                    imagen_encontrada = True
                    break

                except Exception as e:
                    error_msg = f"ID {perro_id} - Fila {row}, Celda {cell.coordinate}: {str(e)}"
                    errores.append(error_msg)
                    print(f"❌ Error: {error_msg}")
                    break

        # Si NO se encontro imagen en ninguna columna de esta fila
        if not imagen_encontrada:
            faltante = {
                'id': perro_id,
                'fila': row,
                'mensaje': f'No se encontró imagen para el ID {perro_id} en la fila {row}'
            }
            imagenes_faltantes.append(faltante)
            print(f"⚠️  ID {perro_id:3d} | Fila {row:3d} | ❌ IMAGEN NO ENCONTRADA - ID SALTADO")
        perro_id += 1

    wb.close()
    print(f"\n{'='*60}")
    print("📊 RESUMEN DEL PROCESO")
    print(f"{'='*60}")
    print(f"✅ Imagenes guardadas: {len(imagenes_guardadas)}")
    print(f"📁 Carpeta destino: {carpeta_salida}")
    print(f"🆔 IDs generados: 1 a {perro_id - 1}")

    if errores:
        print(f"\n⚠️  Errores encontrados: {len(errores)}")
        for error in errores[:5]:
            print(f"   - {error}")
        if len(errores) > 5:
            print(f"   ... y {len(errores) - 5} errores más")

    print(f"{'='*60}\n")

    return {
        'imagenes': imagenes_guardadas,
        'total': len(imagenes_guardadas),
        'errores': errores
    }


def generar_reporte(resultado, archivo_reporte='reporte_imagenes.txt'):
    with open(archivo_reporte, 'w', encoding='utf-8') as f:
        f.write("REPORTE DE EXTRACCIÓN DE IMÁGENES\n")
        f.write("=" * 60 + "\n\n")

        f.write(f"Total de imagenes extraídas: {resultado['total']}\n\n")

        f.write("DETALLE DE IMaGENES:\n")
        f.write("-" * 60 + "\n")
        for img in resultado['imagenes']:
            f.write(f"ID: {img['id']} | Fila: {img['fila']} | Celda: {img['celda']} | Archivo: {img['archivo']}\n")

        if resultado['errores']:
            f.write("\n\nERRORES:\n")
            f.write("-" * 60 + "\n")
            for error in resultado['errores']:
                f.write(f"{error}\n")

    print(f"📄 Reporte guardado en: {archivo_reporte}")


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ARCHIVO_EXCEL = os.path.join(script_dir, "perritos.xlsx")
    CARPETA_SALIDA = os.path.join(script_dir, "imagenes_perros")
    FILA_INICIO = 4
    COLUMNA_IMAGEN = 1  # Columna A (1), B (2), C (3), etc.

    print("=" * 60)
    print("🐕 EXTRACTOR DE IMAGENES DE EXCEL - SISTEMA PATITAS")
    print("=" * 60)
    print()

    if not os.path.exists(ARCHIVO_EXCEL):
        print("❌ Error: No se encuentra el archivo")
        print(f"   Buscando en: {ARCHIVO_EXCEL}")
        print("\n💡 Asegúrate de que el archivo 'perritos.xlsx' esté en la misma carpeta que este script")
        exit(1)

    try:
        resultado = extraer_imagenes_excel(
            ARCHIVO_EXCEL,
            CARPETA_SALIDA,
            fila_inicio=FILA_INICIO,
            columna_imagen=COLUMNA_IMAGEN
        )

        # Generar reporte
        archivo_reporte = os.path.join(script_dir, "reporte_imagenes.txt")
        generar_reporte(resultado, archivo_reporte)

        print("Proceso completado exitosamente")

    except Exception as e:
        print(f"\n❌ Error critico: {e}")
        print("\n💡 Verifica que:")
        print("   1. El archivo Excel no este abierto")
        print("   2. Tengas instaladas las librerías: pip install openpyxl openpyxl-image-loader Pillow")
        exit(1)
