"""
Script para procesar datos de perros desde Excel.
Lee el Excel, valida y transforma los datos según el modelo Django.
Genera un archivo JSON listo para importar a la base de datos.
"""

import openpyxl
import json
import os
import re
from pathlib import Path
from datetime import date


def extraer_seccion(jaula_texto):
    if not jaula_texto:
        return "1"  
    
    # Buscar patrón "Seccion X"
    match = re.search(r'Sección\s+(\d+)', str(jaula_texto), re.IGNORECASE)
    if match:
        return match.group(1)
    return "1"  


def parsear_sexo(sexo_texto):
    if not sexo_texto:
        return None
    
    sexo = str(sexo_texto).strip().lower()
    
    if sexo in ['macho', 'm']:
        return "M"
    elif sexo in ['hembra', 'h']:
        return "H"
    
    return None


def parsear_edad(edad_texto):
    if not edad_texto:
        return (0, 1)  
    
    texto = str(edad_texto).lower().strip()
    
    años = 0
    meses = 0
    
    # Buscar años
    match_años = re.search(r'(\d+)\s*año', texto)
    if match_años:
        años = int(match_años.group(1))
    
    # Buscar meses
    match_meses = re.search(r'(\d+)\s*mes', texto)
    if match_meses:
        meses = int(match_meses.group(1))
    
    # Si solo hay un número sin texto, asumir que son años
    if años == 0 and meses == 0:
        match_numero = re.search(r'^\d+$', texto)
        if match_numero:
            años = int(texto)
    
    # Si ambos siguen en 0, poner 1 mes por defecto
    if años == 0 and meses == 0:
        meses = 1
    
    return (años, meses)


def parsear_tamaño(tamaño_texto):

    if not tamaño_texto:
        return "M"  # Por defecto mediano
    
    texto = str(tamaño_texto).strip().lower()
    
    # Si hay guión, tomar solo la primera parte
    if '-' in texto:
        texto = texto.split('-')[0].strip()
    
    # Mapear tamaños
    if 'pequeñ' in texto or 'chic' in texto or 'small' in texto:
        return "S"
    elif 'median' in texto or 'medium' in texto:
        return "M"
    elif 'grand' in texto or 'large' in texto:
        return "L"
    
    return "M"  # Por defecto mediano


def limpiar_texto(texto):
    """
    Limpia y normaliza texto.
    """
    if not texto:
        return ""
    
    return str(texto).strip()


def procesar_excel_perros(archivo_excel, fila_inicio=4):
    """
    Procesa el archivo Excel y extrae los datos de los perros.
    
    Columnas:
    A: NÚMERO (índice)
    B: JAULA (extraer sección)
    C: NOMBRE
    D: FOTO (no usar)
    E: SEXO
    F: RAZA (no usar)
    G: EDAD
    H: SOCIABLE (no usar)
    I-J: OBSERVACIONES (pueden estar en I o J)
    K: TAMAÑO
    
    Returns:
        list: Lista de diccionarios con los datos de cada perro
    """
    
    print(f"📂 Cargando archivo: {archivo_excel}")
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    sheet = wb.active
    
    perros = []
    errores = []
    
    print(f"🔍 Procesando datos desde fila {fila_inicio}...\n")
    
    for row in range(fila_inicio, sheet.max_row + 1):
        # Leer columnas
        numero = sheet.cell(row, 1).value  # A
        jaula = sheet.cell(row, 2).value   # B
        nombre = sheet.cell(row, 3).value  # C
        # columna D (FOTO) - no usar
        sexo = sheet.cell(row, 5).value    # E
        # columna F (RAZA) - no usar
        edad = sheet.cell(row, 7).value    # G
        # columna H (SOCIABLE) - no usar
        obs_i = sheet.cell(row, 9).value   # I (primera parte observaciones)
        obs_j = sheet.cell(row, 10).value  # J (segunda parte observaciones)
        tamaño = sheet.cell(row, 11).value # K
        
        # Si la fila está vacía (sin número), saltar
        if not numero:
            continue
        
        try:
            # Calcular ID (mismo que el de las imágenes)
            perro_id = row - fila_inicio + 1
            
            # Procesar datos
            seccion = extraer_seccion(jaula)
            nombre_limpio = limpiar_texto(nombre)
            sexo_procesado = parsear_sexo(sexo)
            años, meses = parsear_edad(edad)
            tamaño_procesado = parsear_tamaño(tamaño)
            
            # Combinar observaciones (I y J)
            observaciones = ""
            if obs_i:
                observaciones = limpiar_texto(obs_i)
            if obs_j:
                obs_j_limpio = limpiar_texto(obs_j)
                if obs_j_limpio:
                    observaciones += " " + obs_j_limpio if observaciones else obs_j_limpio
            
            # Validar que tenga sexo (campo obligatorio)
            if not sexo_procesado:
                error_msg = f"ID {perro_id} (Fila {row}): Sexo inválido o vacío '{sexo}'"
                errores.append(error_msg)
                print(f"⚠️  {error_msg}")
                continue
            
            # Crear diccionario con los datos
            perro = {
                'id': perro_id,
                'fila_excel': row,
                'numero_original': numero,
                'nombre': nombre_limpio,
                'seccion': seccion,
                'sexo': sexo_procesado,
                'edad_años': años,
                'edad_meses': meses,
                'tamaño': tamaño_procesado,
                'descripcion': observaciones,
                'fecha_llegada': date.today().isoformat(),
                'estado_adopcion': False,
                'imagen_local': f"perro_id_{perro_id}.png"
            }
            
            perros.append(perro)
            
            print(f"✅ ID {perro_id:3d} | Fila {row:3d} | {nombre_limpio or '(Sin nombre)'} | {sexo_procesado} | {años}a {meses}m | {tamaño_procesado}")
            
        except Exception as e:
            error_msg = f"ID {perro_id} (Fila {row}): Error al procesar - {str(e)}"
            errores.append(error_msg)
            print(f"❌ {error_msg}")
    
    wb.close()
    
    return perros, errores


def guardar_json(perros, archivo_salida='perros_procesados.json'):
    """
    Guarda los datos procesados en un archivo JSON.
    """
    with open(archivo_salida, 'w', encoding='utf-8') as f:
        json.dump(perros, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Datos guardados en: {archivo_salida}")


def generar_reporte(perros, errores, archivo_reporte='reporte_procesamiento.txt'):
    """
    Genera un reporte detallado del procesamiento.
    """
    with open(archivo_reporte, 'w', encoding='utf-8') as f:
        f.write("REPORTE DE PROCESAMIENTO DE DATOS - PERROS\n")
        f.write("=" * 70 + "\n\n")
        
        f.write(f"Total de perros procesados: {len(perros)}\n")
        f.write(f"Total de errores: {len(errores)}\n\n")
        
        # Estadísticas
        machos = sum(1 for p in perros if p['sexo'] == 'M')
        hembras = sum(1 for p in perros if p['sexo'] == 'H')
        
        tamaños = {}
        for p in perros:
            tamaños[p['tamaño']] = tamaños.get(p['tamaño'], 0) + 1
        
        f.write("ESTADÍSTICAS:\n")
        f.write("-" * 70 + "\n")
        f.write(f"Machos: {machos}\n")
        f.write(f"Hembras: {hembras}\n")
        f.write(f"Tamaños: {tamaños}\n\n")
        
        f.write("DETALLE DE PERROS:\n")
        f.write("-" * 70 + "\n")
        for perro in perros:
            f.write(f"ID: {perro['id']} | Fila: {perro['fila_excel']} | "
                   f"Nombre: {perro['nombre'] or '(Sin nombre)'} | "
                   f"Sexo: {perro['sexo']} | "
                   f"Edad: {perro['edad_años']}a {perro['edad_meses']}m | "
                   f"Tamaño: {perro['tamaño']} | "
                   f"Sección: {perro['seccion']}\n")
        
        if errores:
            f.write("\n\nERRORES:\n")
            f.write("-" * 70 + "\n")
            for error in errores:
                f.write(f"{error}\n")
    
    print(f"📄 Reporte guardado en: {archivo_reporte}")


if __name__ == "__main__":
    # Configuración
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ARCHIVO_EXCEL = os.path.join(script_dir, "perritos.xlsx")
    ARCHIVO_JSON = os.path.join(script_dir, "perros_procesados.json")
    ARCHIVO_REPORTE = os.path.join(script_dir, "reporte_procesamiento.txt")
    FILA_INICIO = 4  # Los datos empiezan en fila 4
    
    print("=" * 70)
    print("🐕 PROCESADOR DE DATOS DE EXCEL - SISTEMA PATITAS")
    print("=" * 70)
    print()
    
    # Verificar que el archivo existe
    if not os.path.exists(ARCHIVO_EXCEL):
        print(f"❌ Error: No se encuentra el archivo")
        print(f"   Buscando en: {ARCHIVO_EXCEL}")
        print(f"\n💡 Asegúrate de que el archivo 'perritos.xlsx' esté en la misma carpeta.")
        exit(1)
    
    try:
        # Procesar Excel
        perros, errores = procesar_excel_perros(ARCHIVO_EXCEL, FILA_INICIO)
        
        # Guardar JSON
        guardar_json(perros, ARCHIVO_JSON)
        
        # Generar reporte
        generar_reporte(perros, errores, ARCHIVO_REPORTE)
        
        # Resumen
        print("\n" + "=" * 70)
        print("📊 RESUMEN DEL PROCESO")
        print("=" * 70)
        print(f"✅ Perros procesados: {len(perros)}")
        print(f"⚠️  Errores encontrados: {len(errores)}")
        print(f"💾 Archivo JSON generado: {ARCHIVO_JSON}")
        print(f"📄 Reporte generado: {ARCHIVO_REPORTE}")
        print("=" * 70)
        
        if errores:
            print(f"\n⚠️  Revisa el reporte para ver los detalles de los errores")
        
        print("\n✨ ¡Proceso completado!")
        print("\n📋 Siguiente paso: Revisar el archivo JSON y luego subir a Cloudinary y PostgreSQL")
        
    except Exception as e:
        print(f"\n❌ Error crítico: {e}")
        print(f"\n💡 Verifica que:")
        print("   1. El archivo Excel no esté abierto")
        print("   2. Las columnas estén en el orden correcto")
        print("   3. Tengas instalado: pip install openpyxl")
        exit(1)